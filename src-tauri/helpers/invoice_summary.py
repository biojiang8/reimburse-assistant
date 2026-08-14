#!/usr/bin/env python3
"""发票配套 Excel 生成器 —— 按武汉塞维尔官方订单表模板输出。

结构（单工作表）：
项目名称 / 财务项目码 / 发票号 / 物品名称 / 物品规格 /
供应商编码 / 供应商名称 / 单位 / 单价 / 数量
表头红字加粗（标红字段必填），底部带"注意"说明行。

输入：--json-file 指向的记录 JSON（单个对象或数组）。
字段：project_name / project_code / invoice_number / seller /
items（明细行数组，含 name/spec/unit/quantity/unit_price）。
输出：--output 指向的 .xlsx 文件。
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional


def _as_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _raw_number(invoice_number: str) -> str:
    return invoice_number[4:] if invoice_number.startswith("dzfp") else invoice_number


def write_excel(records: List[Dict[str, Any]], output_path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    headers = [
        "项目名称", "财务项目码", "发票号", "物品名称", "物品规格",
        "供应商编码", "供应商名称", "单位", "单价", "数量",
    ]
    header_font = Font(bold=True, color="FFFF0000")
    gray_font = Font(color="FF333333")
    black_font = Font(color="FF000000")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for record in records:
        raw = _raw_number(str(record.get("invoice_number") or ""))
        items = record.get("items") or []
        if not items:
            items = [{}]
        for item in items:
            values = [
                record.get("project_name") or "",
                record.get("project_code") or "",
                raw,
                item.get("name") or item.get("classification") or "",
                item.get("spec") or "",
                record.get("vendor_code") or "",
                record.get("seller") or "",
                item.get("unit") or "",
                _as_number(item.get("unit_price")),
                _as_number(item.get("quantity")),
            ]
            # 无内容的一律填"无"，不留空格
            values = ["无" if value is None or value == "" else value for value in values]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row, column=column, value=value)
                if column <= 3:
                    cell.font = gray_font
                else:
                    cell.font = black_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center" if column in (3, 8, 9, 10) else "left",
                    vertical="center",
                    wrap_text=(column in (1, 4, 5, 7)),
                )
                # 发票号列设为文本格式，避免科学计数法（无需加撇号）
                if column == 3:
                    cell.number_format = "@"
                if column == 9:
                    cell.number_format = "#,##0.00##"
            row += 1

    note_row = row
    note = sheet.cell(
        row=note_row,
        column=2,
        value=(
            "注意：1、标红字段必填。\n"
            "      2、请误删除表格字段。\n"
            "      3、按发票明细顺序填写。"
        ),
    )
    note.font = Font(color="FF333333")
    note.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [34, 12, 24, 30, 20, 14, 28, 8, 14, 10]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))


def main(argv: List[str]) -> int:
    parser = argparse.ArgumentParser(description="发票 Excel 生成（塞维尔官方模板）")
    # 由 reimburse-helper 的 summary 子命令路由使用，脚本本身忽略
    parser.add_argument("--summary-json", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--json-file", required=True, help="记录 JSON 文件路径")
    parser.add_argument("--output", required=True, help="输出 .xlsx 路径")
    args = parser.parse_args(argv)

    data = json.loads(Path(args.json_file).read_text(encoding="utf-8"))
    if isinstance(data, list):
        records = data
    elif isinstance(data, dict) and isinstance(data.get("records"), list):
        records = data["records"]
    elif isinstance(data, dict):
        records = [data]
    else:
        records = []

    output_path = Path(args.output)
    write_excel(records, output_path)
    print(
        json.dumps(
            {"ok": True, "excel": str(output_path.resolve()), "rows": len(records)},
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
